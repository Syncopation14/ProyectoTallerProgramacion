import openpyxl
wb = openpyxl.load_workbook('bdd/Bases Ficha Colaborador.xlsx', data_only=True)
ws = wb['Sucesion 2024']
## esto extrae las cabeceras de las columnas
Cabecezas = []
for i in range(1,ws.max_column + 1):
    cell_values = ws.cell(row=1,column=i).value
    if cell_values is None:
        break
    Cabecezas.append(cell_values)

### en base al rut debemos obtener los cargos actuales y los cargos a sucesion
def obtenersuce(rut):
    try:
        col_rut = Cabecezas.index('Rut Ocup Actual') + 1
    except ValueError:
        print("Cabecera 'Rut Ocup Actual' no encontrada.")
        return {}

    resultado = {
        "cargo_actual": [],
        "sucesion": []
    }

    for fila in range(2, ws.max_row + 1):
        valor = ws.cell(row=fila, column=col_rut).value
        if valor == rut:
            cargo = ws.cell(row=fila, column=5).value
            sucesion = ws.cell(row=fila, column=20).value
            if cargo:
                resultado["cargo_actual"].append(cargo)
            if sucesion:
                resultado["sucesion"].append(sucesion)

    if not resultado["cargo_actual"] and not resultado["sucesion"]:
       resultado["cargo_actual"].append(4)
       resultado["sucesion"].append(4)

    return resultado
