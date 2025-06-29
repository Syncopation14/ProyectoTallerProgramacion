import openpyxl
# Este script se encarga de cargar un archivo Excel y extraer información de él.
try:
    wb = openpyxl.load_workbook(r"C:\Users\claus\OneDrive\Documentos\cod_claudio\cod_claudio\bdd\Bases Ficha Colaborador.xlsx", data_only=True)
except FileNotFoundError:
    print("Error: El archivo Excel no se encontró en la ruta especificada. Por favor, verifica la ruta.")
    exit()
except Exception as e:
    print(f"Error al cargar el archivo Excel: {e}")
    exit()

try:
    ws = wb['VIG 19_05']
except KeyError:
    print("Error: La hoja 'VIG 19_05' no se encontró en el archivo Excel. Por favor, verifica el nombre de la hoja.")
    exit()

# Esto extrae las cabeceras de las columnas de la primera fila de la hoja.
Cabeceras = []
for i in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=i).value
    if cell_value is None:
        break # Detiene la lectura si la celda de la cabecera está vacía
    Cabeceras.append(cell_value)

print(f"Cabeceras encontradas: {Cabeceras}") # Para depuración

## Con esto se extraen los datos de la columna 'NOMBRE COMPLETO3' y se guardan en un array.
def todoslosnombres():
    datos = []
    # Busca el índice de la columna 'NOMBRE COMPLETO3' en las cabeceras.
    try:
        id_col_nombre = Cabeceras.index('NOMBRE COMPLETO3') + 1 # +1 porque los índices de openpyxl son base 1
    except ValueError:
        print("Error: La cabecera 'NOMBRE COMPLETO3' no se encontró en la hoja 'VIG 19_05'.")
        return []

    # Itera desde la segunda fila (para omitir las cabeceras) hasta el final de la hoja.
    for j in range(2, ws.max_row + 1):
        cell_data = ws.cell(row=j, column=id_col_nombre).value
        if cell_data is None:
            break # Detiene la lectura si la celda de datos está vacía
        datos.append(cell_data)
    return datos

## Esto extrae los datos del trabajador seleccionado.
## 'nombre' es el valor de 'NOMBRE COMPLETO3' del trabajador.
def obtenerDatos(nombre):
    datos = []
    # Definimos las columnas que nos interesan (ejemplo, ajusta según tus necesidades reales de columnas en VIG 19_05)
    # Estas son las columnas que deseas extraer para el trabajador.
    # Por ejemplo, si en 'VIG 19_05' quieres la columna A, B, C, D, E, sus índices serían 1, 2, 3, 4, 5.
    # Es crucial que estos índices se correspondan con las columnas que deseas en tu hoja "VIG 19_05".
    # Usaré los mismos índices que en tu ejemplo (2, 3, 4, 6, 8) pero ten en cuenta que estos corresponden
    # a la numeración de columnas en Excel (columna B, C, D, F, H).
    columnas_a_extraer = [2, 3, 4, 6, 8]

    # Busca el índice de la columna 'NOMBRE COMPLETO3'
    try:
        col_nombres = Cabeceras.index('NOMBRE COMPLETO3') + 1
    except ValueError:
        print("Cabecera 'NOMBRE COMPLETO3' no encontrada en las cabeceras de la hoja actual.")
        return []

    fila_objetivo = None
    # Itera a través de las filas para encontrar el nombre del trabajador.
    for fila in range(2, ws.max_row + 1):
        valor_celda_nombre = ws.cell(row=fila, column=col_nombres).value
        if valor_celda_nombre == nombre:
            fila_objetivo = fila
            break

    if fila_objetivo is None:
        print(f"Advertencia: El nombre '{nombre}' no se encontró en la hoja.")
        return []

    # Extrae los datos de las columnas especificadas para la fila encontrada.
    for col_index in columnas_a_extraer:
        datos.append(ws.cell(row=fila_objetivo, column=col_index).value)
    return datos

# --- Ejemplo de uso (opcional, para probar las funciones) ---
if __name__ == "__main__":
    print("\n--- Probando 'todoslosnombres()' ---")
    nombres_disponibles = todoslosnombres()
    if nombres_disponibles:
        print(f"Primeros 5 nombres disponibles: {nombres_disponibles[:5]}")
        print(f"Total de nombres encontrados: {len(nombres_disponibles)}")

        # Prueba con un nombre si hay alguno disponible
        if nombres_disponibles:
            nombre_ejemplo = nombres_disponibles[0] # Tomamos el primer nombre para el ejemplo
            print(f"\n--- Probando 'obtenerDatos()' con el nombre: '{nombre_ejemplo}' ---")
            datos_trabajador = obtenerDatos(nombre_ejemplo)
            print(f"Datos obtenidos para '{nombre_ejemplo}': {datos_trabajador}")

            # Prueba con un nombre que no existe
            nombre_inexistente = "Nombre No Existente"
            print(f"\n--- Probando 'obtenerDatos()' con el nombre: '{nombre_inexistente}' ---")
            datos_trabajador_inexistente = obtenerDatos(nombre_inexistente)
            print(f"Datos obtenidos para '{nombre_inexistente}': {datos_trabajador_inexistente}")
    else:
        print("No se encontraron nombres para probar.")