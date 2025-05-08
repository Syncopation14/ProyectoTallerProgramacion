import openpyxl
wb = openpyxl.load_workbook('bdd/Bases Ficha Colaborador.xlsx', data_only=True)
ws = wb['Participan']
## esto extrae las cabeceras de las columnas
Cabecezas = []
for i in range(1,ws.max_column + 1):
    cell_values = ws.cell(row=1,column=i).value
    if cell_values is None:
        break
    Cabecezas.append(cell_values)
##con esto se etraen los datos de la fila nombres y se guardan en un array para luego pasarlos
## a la lista desplegable del tk
def todoslosnombres():
    datos = []
    for i in Cabecezas:
        if 'NOMBRE COMPLETO3' == i:
            id = Cabecezas.index(i) + 1  
            for j in range(2, ws.max_row + 1):  
                cell_data = ws.cell(row=j, column=id).value
                if cell_data is None:
                    break
                datos.append(cell_data)
    return datos
## esto extrae los datos del trabajador selecionado del tk.
def obtenerDatos(nombre):
    datos = []
    columnas = [2, 3, 4, 6, 8]

    
    try:
        col_nombres = Cabecezas.index('NOMBRE COMPLETO3') + 1
    except ValueError:
        print("Cabecera 'NOMBRE COMPLETO3' no encontrada.")
        return []

    
    fila_objetivo = None
    for fila in range(2, ws.max_row + 1):  
        valor = ws.cell(row=fila, column=col_nombres).value
        if valor == nombre:
            fila_objetivo = fila
            break

    if fila_objetivo is None:
        print(f"Nombre '{nombre}' no encontrado.")
        return []

    
    datos = [ws.cell(row=fila_objetivo, column=col).value for col in columnas]
    return datos